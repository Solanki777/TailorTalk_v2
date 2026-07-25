"""
Report stage.

Turns the results of the earlier stages (parsed spec, rule engine
findings, AI analysis) into two downloadable artifacts per workspace:

    report.pdf    a readable summary for humans
    report.xlsx   a structured spreadsheet for further filtering/sorting
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.models.workspace import AIAnalysisResult, RuleEngineResult, TestSpec

_SEVERITY_COLORS = {
    "error": colors.HexColor("#DC2626"),
    "warning": colors.HexColor("#D97706"),
    "info": colors.HexColor("#2563EB"),
}


def generate_pdf_report(
    out_path: Path,
    workspace_id: str,
    spec: TestSpec,
    rule_result: RuleEngineResult,
    ai_result: AIAnalysisResult,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=LETTER,
        title="TestPilot AI Report",
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TPTitle", parent=styles["Title"], textColor=colors.HexColor("#312E81"), spaceAfter=2
    )
    meta_style = ParagraphStyle(
        "TPMeta", parent=styles["Normal"], textColor=colors.HexColor("#475569"), fontSize=9
    )
    h2_style = ParagraphStyle(
        "TPHeading2",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#4338CA"),
        spaceBefore=14,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "TPBody", parent=styles["Normal"], fontSize=9, leading=13
    )
    cell_style = ParagraphStyle(
        "TPCell", parent=styles["Normal"], fontSize=8, leading=11
    )
    header_cell_style = ParagraphStyle(
        "TPHeaderCell", parent=cell_style, textColor=colors.white, fontName="Helvetica-Bold"
    )

    def cell(text: str) -> Paragraph:
        return Paragraph(_escape(text), cell_style)

    def header_cell(text: str) -> Paragraph:
        return Paragraph(_escape(text), header_cell_style)

    story = []

    # --- Title block -----------------------------------------------------
    story.append(Paragraph("TestPilot AI — Analysis Report", title_style))
    story.append(Paragraph(f"Workspace ID: {workspace_id}", meta_style))
    story.append(
        Paragraph(f"Test Suite: {spec.test_suite}&nbsp;&nbsp;·&nbsp;&nbsp;Version: {spec.version}", meta_style)
    )
    if spec.base_url:
        story.append(Paragraph(f"Base URL: {spec.base_url}", meta_style))
    story.append(Spacer(1, 2))
    story.append(_hr())

    # --- Summary (labeled 2-column stat grid, not a data table) ----------
    story.append(Paragraph("Summary", h2_style))
    summary_rows = [
        [_stat_label("Total Test Cases"), _stat_label("Rule Engine Errors")],
        [_stat_value(str(len(spec.test_cases))), _stat_value(str(rule_result.error_count), rule_result.error_count > 0)],
        [_stat_label("Rule Engine Warnings"), _stat_label("Rule Engine Info Notes")],
        [_stat_value(str(rule_result.warning_count), rule_result.warning_count > 0), _stat_value(str(rule_result.info_count))],
    ]
    summary_table = Table(summary_rows, colWidths=[2.75 * inch, 2.75 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F5FF")),
                ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#C7D2FE")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E7FF")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.append(summary_table)

    # --- Rule engine findings --------------------------------------------
    story.append(Paragraph("Rule Engine Findings", h2_style))
    if rule_result.findings:
        rows = [[header_cell("Test Case"), header_cell("Rule"), header_cell("Severity"), header_cell("Message")]]
        for f in rule_result.findings:
            rows.append([cell(f.test_case_id or "-"), cell(f.rule), _severity_cell(f.severity.value), cell(f.message)])
        findings_table = Table(
            rows, colWidths=[0.85 * inch, 1.15 * inch, 0.75 * inch, 3.25 * inch], repeatRows=1
        )
        findings_table.setStyle(_table_style(colors.HexColor("#4338CA")))
        story.append(findings_table)
    else:
        story.append(Paragraph("No issues found by the rule engine.", body_style))

    # --- AI analysis -------------------------------------------------------
    story.append(Paragraph("AI Analysis", h2_style))
    if ai_result.skipped_reason:
        story.append(
            Paragraph(f"<i>AI analysis was not available: {_escape(ai_result.skipped_reason)}</i>", body_style)
        )
    else:
        if ai_result.model_used:
            story.append(Paragraph(f"Model: {ai_result.model_used}", meta_style))
            story.append(Spacer(1, 4))
        story.append(Paragraph(_escape(ai_result.summary), body_style))
        story.append(Spacer(1, 8))
        if ai_result.coverage_gaps:
            story.append(Paragraph("Coverage Gaps", styles["Heading4"]))
            for gap in ai_result.coverage_gaps:
                story.append(Paragraph(f"•&nbsp;&nbsp;{_escape(gap)}", body_style))
            story.append(Spacer(1, 8))
        if ai_result.findings:
            rows = [[header_cell("Category"), header_cell("Severity"), header_cell("Message")]]
            for f in ai_result.findings:
                rows.append([cell(f.category), _severity_cell(f.severity.value), cell(f.message)])
            ai_table = Table(rows, colWidths=[1.2 * inch, 0.8 * inch, 4.0 * inch], repeatRows=1)
            ai_table.setStyle(_table_style(colors.HexColor("#7C3AED")))
            story.append(ai_table)

    doc.build(story)


def _table_style(header_bg) -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
        ]
    )


def _severity_cell(severity: str) -> Paragraph:
    color = _SEVERITY_COLORS.get(severity, colors.black)
    style = ParagraphStyle(
        f"Sev_{severity}",
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=color,
    )
    return Paragraph(severity.upper(), style)


def _stat_label(text: str) -> Paragraph:
    style = ParagraphStyle("StatLabel", fontSize=8.5, textColor=colors.HexColor("#4338CA"), fontName="Helvetica-Bold")
    return Paragraph(text, style)


def _stat_value(text: str, alert: bool = False) -> Paragraph:
    color = colors.HexColor("#DC2626") if alert else colors.HexColor("#111827")
    style = ParagraphStyle("StatValue", fontSize=16, textColor=color, fontName="Helvetica-Bold")
    return Paragraph(text, style)


def _hr():
    line = Table([[""]], colWidths=[7 * inch], rowHeights=[1])
    line.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, -1), 0.75, colors.HexColor("#E0E7FF"))]))
    return line


def _escape(text: str) -> str:
    """Escape text for safe use inside a ReportLab Paragraph (which parses a mini-HTML subset)."""
    return (
        (text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def generate_xlsx_report(
    out_path: Path,
    workspace_id: str,
    spec: TestSpec,
    rule_result: RuleEngineResult,
    ai_result: AIAnalysisResult,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    # --- Sheet 1: Test Cases ---------------------------------------------
    ws1 = wb.active
    ws1.title = "Test Cases"
    headers = ["ID", "Name", "Method", "Endpoint", "Expected Status", "Has Request Body", "Has Expected Response"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
    for case in spec.test_cases:
        ws1.append(
            [
                case.id,
                case.name,
                case.method or "",
                case.endpoint or "",
                case.expected_status if case.expected_status is not None else "",
                "Yes" if case.request_body else "No",
                "Yes" if case.expected_response else "No",
            ]
        )
    _autosize(ws1, headers)

    # --- Sheet 2: Rule Engine Findings ------------------------------------
    ws2 = wb.create_sheet("Rule Engine Findings")
    headers2 = ["Test Case", "Rule", "Severity", "Message"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for f in rule_result.findings:
        ws2.append([f.test_case_id or "-", f.rule, f.severity.value.upper(), f.message])
        ws2.cell(ws2.max_row, 4).alignment = wrap
    _autosize(ws2, headers2)

    # --- Sheet 3: AI Analysis ---------------------------------------------
    ws3 = wb.create_sheet("AI Analysis")
    ws3.append(["Summary"])
    ws3["A1"].font = Font(bold=True)
    ws3.append([ai_result.skipped_reason and f"Skipped: {ai_result.skipped_reason}" or ai_result.summary])
    ws3.append([])
    if ai_result.coverage_gaps:
        ws3.append(["Coverage Gaps"])
        ws3.cell(ws3.max_row, 1).font = Font(bold=True)
        for gap in ai_result.coverage_gaps:
            ws3.append([gap])
        ws3.append([])
    if ai_result.findings:
        headers3 = ["Category", "Severity", "Message"]
        ws3.append(headers3)
        for cell in ws3[ws3.max_row]:
            cell.fill = header_fill
            cell.font = header_font
        for f in ai_result.findings:
            ws3.append([f.category, f.severity.value.upper(), f.message])
            ws3.cell(ws3.max_row, 3).alignment = wrap
    ws3.column_dimensions["A"].width = 90

    wb.save(out_path)


def _autosize(ws, headers: list[str], min_width: int = 12, max_width: int = 50) -> None:
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        longest = max(
            [len(header)]
            + [len(str(cell.value)) for cell in ws[col_letter] if cell.value is not None]
        )
        ws.column_dimensions[col_letter].width = min(max(longest + 2, min_width), max_width)